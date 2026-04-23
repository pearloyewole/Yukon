#!/usr/bin/env python3
"""Build a browser-ready river package JSON from shapefile + overview workbook + MAT files.

Supports two modes:
- Explicit inputs (shp/dbf/shx + overview + mat glob)
- Zip bundle input (auto-detects components)
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import glob
import io
import json
import math
import random
import tempfile
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import scipy.io as sio

try:
    import shapefile  # pyshp
except ImportError as exc:
    raise SystemExit(
        "Missing dependency 'pyshp'. Install with: python -m pip install pyshp"
    ) from exc

try:
    import tifffile
except ImportError:
    tifffile = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a single JSON package for the Three.js river viewer from a river dataset."
        )
    )
    parser.add_argument("--river-id", required=True, help="Short river identifier, e.g. huslia")
    parser.add_argument("--input-zip", help="Path to zip containing shapefile + overview xlsx + mat files")
    parser.add_argument("--shp", help="Path to .shp")
    parser.add_argument("--dbf", help="Path to .dbf (optional if same basename as .shp)")
    parser.add_argument("--shx", help="Path to .shx (optional if same basename as .shp)")
    parser.add_argument("--overview", help="Path to overview .xlsx")
    parser.add_argument("--mat-glob", default="data/mat/*.mat", help="Glob pattern for mat files")
    parser.add_argument(
        "--out",
        default=None,
        help="Output JSON path (default: public/river-packages/<river-id>.json)",
    )
    parser.add_argument(
        "--velocity-key",
        default="xsGridQs",
        help="MAT field for velocity grid",
    )
    parser.add_argument(
        "--mask-key",
        default="mask_temp",
        help="MAT field for mask grid",
    )
    parser.add_argument(
        "--sample-grid-size",
        type=int,
        default=24,
        help="Downsampled square size for velocity preview data",
    )
    parser.add_argument(
        "--sonar-zip",
        action="append",
        default=[],
        help=(
            "Path to a zip containing sonar bottom CSV files "
            "(repeat flag to provide multiple zips)"
        ),
    )
    parser.add_argument(
        "--sonar-csv-glob",
        action="append",
        default=[],
        help=(
            "Glob pattern for sonar bottom CSV files in explicit mode "
            "(repeat flag for multiple patterns)"
        ),
    )
    parser.add_argument(
        "--sonar-max-points",
        type=int,
        default=120000,
        help="Max number of sonar bottom points to store in package JSON",
    )
    parser.add_argument(
        "--sonar-min-depth-m",
        type=float,
        default=0.0,
        help="Minimum sonar depth (meters) to include",
    )
    parser.add_argument(
        "--elevation-tif",
        action="append",
        default=[],
        help=(
            "Path to an elevation GeoTIFF (.tif/.tiff) "
            "(repeat flag to provide multiple rasters)"
        ),
    )
    parser.add_argument(
        "--elevation-max-grid",
        type=int,
        default=240,
        help="Maximum rows/cols to retain for sampled elevation grid output",
    )
    return parser.parse_args()


def find_first(root: Path, pattern: str) -> Path | None:
    matches = sorted(root.rglob(pattern))
    return matches[0] if matches else None


def resolve_inputs(
    args: argparse.Namespace,
) -> tuple[Path, Path, Path, Path, list[Path], dict[str, Any], Path | None]:
    source: dict[str, Any] = {}

    if args.input_zip:
        input_zip = Path(args.input_zip).expanduser().resolve()
        if not input_zip.exists():
            raise SystemExit(f"Zip not found: {input_zip}")

        temp_dir = Path(tempfile.mkdtemp(prefix="river_pkg_"))
        with zipfile.ZipFile(input_zip, "r") as zf:
            zf.extractall(temp_dir)

        shp = find_first(temp_dir, "*.shp")
        dbf = find_first(temp_dir, "*.dbf")
        shx = find_first(temp_dir, "*.shx")
        overview = find_first(temp_dir, "*.xlsx")
        mat_files = sorted(temp_dir.rglob("*.mat"))

        if not shp or not dbf or not shx or not overview or not mat_files:
            raise SystemExit(
                "Zip must contain .shp/.dbf/.shx + overview .xlsx + one or more .mat files"
            )

        source.update(
            {
                "input_zip": str(input_zip),
                "extracted_dir": str(temp_dir),
            }
        )
        return shp, dbf, shx, overview, mat_files, source, temp_dir

    if not args.shp or not args.overview:
        raise SystemExit("Provide --input-zip OR explicit --shp and --overview inputs")

    shp = Path(args.shp).expanduser().resolve()
    if not shp.exists():
        raise SystemExit(f"Shapefile not found: {shp}")

    dbf = Path(args.dbf).expanduser().resolve() if args.dbf else shp.with_suffix(".dbf")
    shx = Path(args.shx).expanduser().resolve() if args.shx else shp.with_suffix(".shx")
    overview = Path(args.overview).expanduser().resolve()

    for p in [dbf, shx, overview]:
        if not p.exists():
            raise SystemExit(f"Required file not found: {p}")

    mat_files = [Path(p).resolve() for p in sorted(glob.glob(args.mat_glob))]
    if not mat_files:
        raise SystemExit(f"No MAT files matched: {args.mat_glob}")

    source.update(
        {
            "shp": str(shp),
            "dbf": str(dbf),
            "shx": str(shx),
            "overview": str(overview),
            "mat_glob": args.mat_glob,
        }
    )
    return shp, dbf, shx, overview, mat_files, source, None


def resolve_sonar_inputs(
    args: argparse.Namespace,
    extracted_dir: Path | None,
    source: dict[str, Any],
) -> tuple[list[Path], list[Path]]:
    sonar_zip_paths: list[Path] = []
    sonar_csv_paths: list[Path] = []

    for zip_path_str in args.sonar_zip:
        zip_path = Path(zip_path_str).expanduser().resolve()
        if not zip_path.exists():
            raise SystemExit(f"Sonar zip not found: {zip_path}")
        sonar_zip_paths.append(zip_path)

    for pattern in args.sonar_csv_glob:
        matches = [Path(p).resolve() for p in sorted(glob.glob(pattern))]
        sonar_csv_paths.extend(matches)

    if extracted_dir is not None:
        # When the main dataset arrives as a zip, auto-scan any CSV files inside.
        sonar_csv_paths.extend(sorted(extracted_dir.rglob("*.csv")))

    # Preserve order while de-duplicating.
    unique_zip_paths = list(dict.fromkeys(sonar_zip_paths))
    unique_csv_paths = list(dict.fromkeys(sonar_csv_paths))

    if unique_zip_paths:
        source["sonar_zip"] = [str(p) for p in unique_zip_paths]
    if args.sonar_csv_glob:
        source["sonar_csv_glob"] = args.sonar_csv_glob
    if extracted_dir is not None:
        source["input_zip_csv_candidates"] = len(unique_csv_paths)

    return unique_zip_paths, unique_csv_paths


def resolve_elevation_inputs(
    args: argparse.Namespace,
    extracted_dir: Path | None,
    source: dict[str, Any],
) -> list[Path]:
    tif_paths: list[Path] = []
    for tif_path_str in args.elevation_tif:
        tif_path = Path(tif_path_str).expanduser().resolve()
        if not tif_path.exists():
            raise SystemExit(f"Elevation TIFF not found: {tif_path}")
        if tif_path.suffix.lower() not in (".tif", ".tiff"):
            raise SystemExit(f"Elevation input must be .tif/.tiff: {tif_path}")
        tif_paths.append(tif_path)

    if extracted_dir is not None:
        tif_paths.extend(sorted(extracted_dir.rglob("*.tif")))
        tif_paths.extend(sorted(extracted_dir.rglob("*.tiff")))

    filtered_tif_paths: list[Path] = []
    for tif_path in tif_paths:
        parts_lower = {part.lower() for part in tif_path.parts}
        if "__macosx" in parts_lower or tif_path.name.startswith("."):
            continue
        filtered_tif_paths.append(tif_path)

    unique_tif_paths = list(dict.fromkeys(filtered_tif_paths))
    if unique_tif_paths:
        source["elevation_tif"] = [str(p) for p in unique_tif_paths]
    return unique_tif_paths


def parse_nodata_value(raw_value: Any) -> float | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, bytes):
        raw_text = raw_value.decode("utf-8", errors="ignore")
    else:
        raw_text = str(raw_value)
    text = raw_text.strip().strip("\x00")
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    return value if math.isfinite(value) else None


def extract_geotiff_reference(page: Any, rows: int, cols: int) -> dict[str, Any] | None:
    tags = page.tags
    x_origin = None
    y_origin = None
    pixel_size_x = None
    pixel_size_y = None
    georef_mode = None

    transform_tag = tags.get("ModelTransformationTag")
    if transform_tag is not None:
        vals = tuple(float(v) for v in transform_tag.value)
        if len(vals) >= 16:
            # GeoTIFF 4x4 transform matrix in row-major order.
            # X = m0*col + m1*row + m3
            # Y = m4*col + m5*row + m7
            x_origin = vals[3]
            y_origin = vals[7]
            pixel_size_x = vals[0]
            pixel_size_y = vals[5]
            georef_mode = "model_transformation"

    if (
        x_origin is None
        or y_origin is None
        or pixel_size_x is None
        or pixel_size_y is None
    ):
        scale_tag = tags.get("ModelPixelScaleTag")
        tie_tag = tags.get("ModelTiepointTag")
        if scale_tag is None or tie_tag is None:
            return None
        scale = tuple(float(v) for v in scale_tag.value)
        tie = tuple(float(v) for v in tie_tag.value)
        if len(scale) < 2 or len(tie) < 6:
            return None

        i0, j0, _, x0, y0, _ = tie[:6]
        sx = scale[0]
        sy = scale[1]
        if not (math.isfinite(sx) and math.isfinite(sy) and sx != 0 and sy != 0):
            return None

        # North-up GeoTIFF convention: row index increases downward, so model Y decreases.
        pixel_size_x = sx
        pixel_size_y = -abs(sy)
        x_origin = x0 - i0 * pixel_size_x
        y_origin = y0 - j0 * pixel_size_y
        georef_mode = "tiepoint_scale"

    if not all(
        math.isfinite(v)
        for v in (
            x_origin,
            y_origin,
            pixel_size_x,
            pixel_size_y,
        )
    ):
        return None

    x_last = x_origin + (cols - 1) * pixel_size_x
    y_last = y_origin + (rows - 1) * pixel_size_y
    bbox_min_x = min(x_origin, x_last)
    bbox_max_x = max(x_origin, x_last)
    bbox_min_y = min(y_origin, y_last)
    bbox_max_y = max(y_origin, y_last)

    ascii_tag = tags.get("GeoAsciiParamsTag")
    crs_name = None
    if ascii_tag is not None:
        raw_crs = str(ascii_tag.value).replace("|", " ").strip()
        crs_name = " ".join(raw_crs.split()) if raw_crs else None

    nodata_tag = tags.get("GDAL_NODATA")
    nodata = parse_nodata_value(nodata_tag.value) if nodata_tag is not None else None

    return {
        "mode": georef_mode,
        "x_origin": float(x_origin),
        "y_origin": float(y_origin),
        "pixel_size_x": float(pixel_size_x),
        "pixel_size_y": float(pixel_size_y),
        "bbox": {
            "min_x": float(bbox_min_x),
            "max_x": float(bbox_max_x),
            "min_y": float(bbox_min_y),
            "max_y": float(bbox_max_y),
        },
        "crs_name": crs_name,
        "nodata": nodata,
    }


def clip_index_range_to_bbox(
    *,
    rows: int,
    cols: int,
    x_origin: float,
    y_origin: float,
    pixel_size_x: float,
    pixel_size_y: float,
    clip_bbox: dict[str, float] | None,
) -> tuple[int, int, int, int]:
    row_start = 0
    row_end = rows - 1
    col_start = 0
    col_end = cols - 1

    if not clip_bbox:
        return row_start, row_end, col_start, col_end

    min_x = clip_bbox.get("min_x")
    max_x = clip_bbox.get("max_x")
    min_y = clip_bbox.get("min_y")
    max_y = clip_bbox.get("max_y")
    if not all(
        isinstance(v, (int, float)) and math.isfinite(float(v))
        for v in (min_x, max_x, min_y, max_y)
    ):
        return row_start, row_end, col_start, col_end

    x_candidates = [
        (float(min_x) - x_origin) / pixel_size_x,
        (float(max_x) - x_origin) / pixel_size_x,
    ]
    y_candidates = [
        (float(min_y) - y_origin) / pixel_size_y,
        (float(max_y) - y_origin) / pixel_size_y,
    ]

    raw_col_start = math.floor(min(x_candidates))
    raw_col_end = math.ceil(max(x_candidates))
    raw_row_start = math.floor(min(y_candidates))
    raw_row_end = math.ceil(max(y_candidates))

    col_start = max(0, min(cols - 1, raw_col_start))
    col_end = max(0, min(cols - 1, raw_col_end))
    row_start = max(0, min(rows - 1, raw_row_start))
    row_end = max(0, min(rows - 1, raw_row_end))

    if col_end < col_start:
        col_start, col_end = col_end, col_start
    if row_end < row_start:
        row_start, row_end = row_end, row_start

    # Keep a minimum footprint to avoid degenerate meshes.
    if col_end - col_start < 8:
        col_start = 0
        col_end = cols - 1
    if row_end - row_start < 8:
        row_start = 0
        row_end = rows - 1

    return row_start, row_end, col_start, col_end


def read_elevation_raster(
    elevation_tif_path: Path,
    *,
    max_grid: int,
    river_bbox: dict[str, Any] | None,
) -> dict[str, Any] | None:
    if tifffile is None:
        raise SystemExit(
            "Missing dependency 'tifffile'. Install with: python -m pip install tifffile"
        )

    with tifffile.TiffFile(elevation_tif_path) as tif:
        if len(tif.pages) == 0:
            return None

        page = tif.pages[0]
        raw = page.asarray(out="memmap")
        if raw.ndim == 2:
            raster = raw
        elif raw.ndim == 3:
            # If multi-band, default to first band.
            if raw.shape[0] <= raw.shape[-1]:
                raster = raw[0]
            else:
                raster = raw[..., 0]
        else:
            return None

        raster = np.asarray(raster)
        if raster.ndim != 2:
            return None

        rows, cols = raster.shape
        georef = extract_geotiff_reference(page, rows, cols)
        if georef is None:
            return None

        clip_bbox = None
        if river_bbox is not None:
            rb_min_x = to_float(river_bbox.get("min_x"))
            rb_max_x = to_float(river_bbox.get("max_x"))
            rb_min_y = to_float(river_bbox.get("min_y"))
            rb_max_y = to_float(river_bbox.get("max_y"))
            if None not in (rb_min_x, rb_max_x, rb_min_y, rb_max_y):
                span_x = max(1.0, rb_max_x - rb_min_x)
                span_y = max(1.0, rb_max_y - rb_min_y)
                margin = max(800.0, 0.2 * max(span_x, span_y))
                clip_bbox = {
                    "min_x": rb_min_x - margin,
                    "max_x": rb_max_x + margin,
                    "min_y": rb_min_y - margin,
                    "max_y": rb_max_y + margin,
                }

        row_start, row_end, col_start, col_end = clip_index_range_to_bbox(
            rows=rows,
            cols=cols,
            x_origin=georef["x_origin"],
            y_origin=georef["y_origin"],
            pixel_size_x=georef["pixel_size_x"],
            pixel_size_y=georef["pixel_size_y"],
            clip_bbox=clip_bbox,
        )

        max_grid = max(16, int(max_grid))
        sample_rows = min(max_grid, row_end - row_start + 1)
        sample_cols = min(max_grid, col_end - col_start + 1)

        row_indices = np.linspace(row_start, row_end, sample_rows).astype(int)
        col_indices = np.linspace(col_start, col_end, sample_cols).astype(int)
        row_indices = np.unique(row_indices)
        col_indices = np.unique(col_indices)
        if row_indices.size < 2 or col_indices.size < 2:
            return None

        sampled = np.asarray(raster[np.ix_(row_indices, col_indices)], dtype=float)
        nodata = georef.get("nodata")

        invalid = ~np.isfinite(sampled)
        if nodata is not None:
            nodata_tol = max(1e-6, abs(float(nodata)) * 1e-6)
            invalid |= np.isclose(sampled, float(nodata), rtol=0.0, atol=nodata_tol)
        sampled[invalid] = np.nan

        valid = sampled[np.isfinite(sampled)]
        if valid.size == 0:
            return None

        values: list[list[float | None]] = []
        for r in range(sampled.shape[0]):
            row_values: list[float | None] = []
            for c in range(sampled.shape[1]):
                value = sampled[r, c]
                row_values.append(float(value) if math.isfinite(value) else None)
            values.append(row_values)

    sample_min_x = georef["x_origin"] + float(np.min(col_indices)) * georef["pixel_size_x"]
    sample_max_x = georef["x_origin"] + float(np.max(col_indices)) * georef["pixel_size_x"]
    sample_min_y = georef["y_origin"] + float(np.min(row_indices)) * georef["pixel_size_y"]
    sample_max_y = georef["y_origin"] + float(np.max(row_indices)) * georef["pixel_size_y"]

    elev_p02 = float(np.percentile(valid, 2))
    elev_p50 = float(np.percentile(valid, 50))
    elev_p98 = float(np.percentile(valid, 98))
    relief = max(0.001, elev_p98 - elev_p02)
    vertical_scale = max(0.2, min(8.0, 12.0 / relief))

    return {
        "source_file": str(elevation_tif_path),
        "shape": {
            "rows": int(rows),
            "cols": int(cols),
        },
        "bbox": {
            "min_x": float(min(sample_min_x, sample_max_x)),
            "max_x": float(max(sample_min_x, sample_max_x)),
            "min_y": float(min(sample_min_y, sample_max_y)),
            "max_y": float(max(sample_min_y, sample_max_y)),
        },
        "georeference": {
            "mode": georef["mode"],
            "x_origin": georef["x_origin"],
            "y_origin": georef["y_origin"],
            "pixel_size_x": georef["pixel_size_x"],
            "pixel_size_y": georef["pixel_size_y"],
            "crs_name": georef["crs_name"],
            "nodata": nodata,
        },
        "sample": {
            "rows": int(row_indices.size),
            "cols": int(col_indices.size),
            "row_indices": [int(v) for v in row_indices.tolist()],
            "col_indices": [int(v) for v in col_indices.tolist()],
            "values": values,
            "valid_count": int(valid.size),
            "value_stats": {
                "min": float(np.min(valid)),
                "max": float(np.max(valid)),
                "mean": float(np.mean(valid)),
            },
        },
        "display": {
            "elevation_reference_m": elev_p50,
            "vertical_scale": vertical_scale,
            "clip_percentile_low_m": elev_p02,
            "clip_percentile_high_m": elev_p98,
        },
    }


def load_elevation_raster_data(
    *,
    elevation_tif_paths: list[Path],
    max_grid: int,
    river_bbox: dict[str, Any] | None,
) -> dict[str, Any] | None:
    if not elevation_tif_paths:
        return None

    for tif_path in elevation_tif_paths:
        raster = read_elevation_raster(
            tif_path,
            max_grid=max_grid,
            river_bbox=river_bbox,
        )
        if raster is not None:
            return raster

    return None


def to_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.floating, float)):
        val = float(value)
        return val if math.isfinite(val) else None
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def normalize_header(name: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")


def to_float(value: Any) -> float | None:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(num):
        return None
    return num


def normalize_latitude(value: Any) -> float | None:
    lat = to_float(value)
    if lat is None:
        return None

    while abs(lat) > 90 and abs(lat) > 1:
        lat /= 10.0

    if abs(lat) > 90:
        return None
    return lat


def normalize_longitude(value: Any, reference_longitude: float | None) -> float | None:
    lon = to_float(value)
    if lon is None:
        return None

    while abs(lon) > 180 and abs(lon) > 1:
        lon /= 10.0

    if abs(lon) > 180:
        return None

    # Fix occasional sign flips (e.g., +156 instead of -156) by matching
    # the dominant hemisphere seen in the overview table.
    if reference_longitude is not None and abs(lon) > 90:
        if reference_longitude < 0 and lon > 0:
            lon = -lon
        elif reference_longitude > 0 and lon < 0:
            lon = -lon

    return lon


def overview_longitude_reference(overview_rows: list[dict[str, Any]]) -> float | None:
    raw_longitudes: list[float] = []
    for row in overview_rows:
        for key in ("start_longitude", "end_longitude"):
            lon = to_float(row.get(key))
            if lon is None:
                continue
            if abs(lon) <= 180:
                raw_longitudes.append(lon)

    if not raw_longitudes:
        return None
    return float(np.median(raw_longitudes))


def build_lonlat_to_xy_transform(overview_rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    control_points: list[tuple[float, float, float, float]] = []
    reference_longitude = overview_longitude_reference(overview_rows)

    for row in overview_rows:
        candidates = [
            (
                row.get("start_longitude"),
                row.get("start_latitude"),
                row.get("cross_section_start_x__utm"),
                row.get("cross_section_start_y__utm"),
            ),
            (
                row.get("end_longitude"),
                row.get("end_latitude"),
                row.get("cross_section_end_x__utm"),
                row.get("cross_section_end_y__utm"),
            ),
        ]

        for lon_val, lat_val, x_val, y_val in candidates:
            lon = normalize_longitude(lon_val, reference_longitude)
            lat = normalize_latitude(lat_val)
            x = to_float(x_val)
            y = to_float(y_val)
            if None in (lon, lat, x, y):
                continue

            if not (0 <= x <= 1_000_000 and 0 <= y <= 10_000_000):
                continue

            control_points.append((lon, lat, x, y))

    if len(control_points) < 3:
        return None

    ll = np.array([[lon, lat, 1.0] for lon, lat, _, _ in control_points], dtype=float)
    xs = np.array([x for _, _, x, _ in control_points], dtype=float)
    ys = np.array([y for _, _, _, y in control_points], dtype=float)

    x_coeff, _, _, _ = np.linalg.lstsq(ll, xs, rcond=None)
    y_coeff, _, _, _ = np.linalg.lstsq(ll, ys, rcond=None)

    pred_x = ll @ x_coeff
    pred_y = ll @ y_coeff
    rmse_x = float(np.sqrt(np.mean((pred_x - xs) ** 2)))
    rmse_y = float(np.sqrt(np.mean((pred_y - ys) ** 2)))

    return {
        "type": "affine_lonlat_to_xy",
        "control_points": len(control_points),
        "coefficients": {
            "x": [float(v) for v in x_coeff],
            "y": [float(v) for v in y_coeff],
        },
        "fit_rmse_m": {
            "x": rmse_x,
            "y": rmse_y,
        },
        "reference_longitude": reference_longitude,
    }


def transform_lonlat_to_xy(transform: dict[str, Any], lon: float, lat: float) -> tuple[float, float]:
    x_coeff = transform["coefficients"]["x"]
    y_coeff = transform["coefficients"]["y"]
    x = x_coeff[0] * lon + x_coeff[1] * lat + x_coeff[2]
    y = y_coeff[0] * lon + y_coeff[1] * lat + y_coeff[2]
    return float(x), float(y)


def match_sonar_columns(fieldnames: list[str]) -> tuple[str, str, str] | None:
    normalized = {normalize_header(name): name for name in fieldnames if name is not None}

    lat_col = None
    lon_col = None
    depth_col = None

    for key in ("latitude", "lat"):
        if key in normalized:
            lat_col = normalized[key]
            break

    for key in ("longitude", "long", "lon"):
        if key in normalized:
            lon_col = normalized[key]
            break

    for key in ("depth_m", "depth__m", "depth"):
        if key in normalized:
            depth_col = normalized[key]
            break

    if depth_col is None:
        for norm_name, raw_name in normalized.items():
            if norm_name.startswith("depth"):
                depth_col = raw_name
                break

    if not lat_col or not lon_col or not depth_col:
        return None
    return lat_col, lon_col, depth_col


def summarize_sonar_csv_stream(
    stream: io.TextIOBase,
    label: str,
    *,
    transform: dict[str, Any],
    min_depth_m: float,
    max_points: int,
    rng: random.Random,
    reservoir: list[tuple[float, float, float, int]],
    global_state: dict[str, Any],
    file_index: int,
) -> dict[str, Any]:
    reader = csv.DictReader(stream)
    if not reader.fieldnames:
        return {"name": label, "rows": 0, "valid_points": 0, "skipped": "missing_header"}

    col_match = match_sonar_columns(reader.fieldnames)
    if col_match is None:
        return {
            "name": label,
            "rows": 0,
            "valid_points": 0,
            "skipped": "missing_lat_lon_depth_columns",
        }

    lat_col, lon_col, depth_col = col_match
    ref_lon = to_float(transform.get("reference_longitude"))

    rows = 0
    valid_points = 0
    depth_min = math.inf
    depth_max = -math.inf
    depth_sum = 0.0

    for row in reader:
        rows += 1

        lat = normalize_latitude(row.get(lat_col))
        lon = normalize_longitude(row.get(lon_col), ref_lon)
        depth = to_float(row.get(depth_col))
        if None in (lat, lon, depth):
            continue
        if depth < min_depth_m:
            continue

        x, y = transform_lonlat_to_xy(transform, lon, lat)

        valid_points += 1
        global_state["valid_points"] += 1
        global_state["depth_sum"] += depth
        global_state["depth_min"] = min(global_state["depth_min"], depth)
        global_state["depth_max"] = max(global_state["depth_max"], depth)
        global_state["min_x"] = min(global_state["min_x"], x)
        global_state["max_x"] = max(global_state["max_x"], x)
        global_state["min_y"] = min(global_state["min_y"], y)
        global_state["max_y"] = max(global_state["max_y"], y)

        depth_min = min(depth_min, depth)
        depth_max = max(depth_max, depth)
        depth_sum += depth

        if max_points <= 0:
            continue

        point = (x, y, depth, file_index)
        if len(reservoir) < max_points:
            reservoir.append(point)
        else:
            j = rng.randrange(global_state["valid_points"])
            if j < max_points:
                reservoir[j] = point

    return {
        "index": file_index,
        "name": label,
        "rows": rows,
        "valid_points": valid_points,
        "depth_m": {
            "min": depth_min if valid_points else None,
            "max": depth_max if valid_points else None,
            "mean": (depth_sum / valid_points) if valid_points else None,
        },
    }


def load_sonar_bottom_data(
    *,
    overview_rows: list[dict[str, Any]],
    sonar_zip_paths: list[Path],
    sonar_csv_paths: list[Path],
    max_points: int,
    min_depth_m: float,
) -> dict[str, Any] | None:
    if not sonar_zip_paths and not sonar_csv_paths:
        return None

    transform = build_lonlat_to_xy_transform(overview_rows)
    if transform is None:
        raise SystemExit(
            "Unable to derive lon/lat -> river XY transform from overview workbook. "
            "Need at least 3 records with both lat/lon and UTM start/end coordinates."
        )

    rng = random.Random(0)
    reservoir: list[tuple[float, float, float, int]] = []
    global_state = {
        "valid_points": 0,
        "depth_sum": 0.0,
        "depth_min": math.inf,
        "depth_max": -math.inf,
        "min_x": math.inf,
        "max_x": -math.inf,
        "min_y": math.inf,
        "max_y": -math.inf,
    }

    file_summaries: list[dict[str, Any]] = []
    file_index = 0

    for csv_path in sonar_csv_paths:
        if not csv_path.exists():
            continue
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
            file_summaries.append(
                summarize_sonar_csv_stream(
                    stream,
                    str(csv_path),
                    transform=transform,
                    min_depth_m=min_depth_m,
                    max_points=max_points,
                    rng=rng,
                    reservoir=reservoir,
                    global_state=global_state,
                    file_index=file_index,
                )
            )
        file_index += 1

    for sonar_zip in sonar_zip_paths:
        with zipfile.ZipFile(sonar_zip, "r") as zf:
            for info in sorted(zf.infolist(), key=lambda item: item.filename):
                if info.is_dir() or not info.filename.lower().endswith(".csv"):
                    continue

                with zf.open(info, "r") as raw:
                    with io.TextIOWrapper(
                        raw, encoding="utf-8-sig", errors="replace", newline=""
                    ) as stream:
                        label = f"{sonar_zip.name}:{info.filename}"
                        file_summaries.append(
                            summarize_sonar_csv_stream(
                                stream,
                                label,
                                transform=transform,
                                min_depth_m=min_depth_m,
                                max_points=max_points,
                                rng=rng,
                                reservoir=reservoir,
                                global_state=global_state,
                                file_index=file_index,
                            )
                        )
                file_index += 1

    valid_files = [f for f in file_summaries if f.get("valid_points", 0) > 0]
    if global_state["valid_points"] == 0:
        return None

    sampled_by_file = Counter(point[3] for point in reservoir)
    for summary in file_summaries:
        idx = summary.pop("index", None)
        summary["sampled_points"] = sampled_by_file.get(idx, 0) if idx is not None else 0

    points = [[float(x), float(y), float(depth)] for x, y, depth, _ in reservoir]

    return {
        "transform": transform,
        "point_fields": ["x", "y", "depth_m"],
        "points": points,
        "point_count_input": int(global_state["valid_points"]),
        "point_count_sampled": len(points),
        "sample_limit": int(max_points),
        "min_depth_m": float(min_depth_m),
        "depth_m": {
            "min": float(global_state["depth_min"]),
            "max": float(global_state["depth_max"]),
            "mean": float(global_state["depth_sum"] / global_state["valid_points"]),
        },
        "bbox": {
            "min_x": float(global_state["min_x"]),
            "max_x": float(global_state["max_x"]),
            "min_y": float(global_state["min_y"]),
            "max_y": float(global_state["max_y"]),
        },
        "file_count": len(valid_files),
        "files": file_summaries,
    }


def load_shapefile_points(shp: Path, dbf: Path, shx: Path) -> dict[str, Any]:
    reader = shapefile.Reader(str(shp), dbf=str(dbf), shx=str(shx))
    field_defs = reader.fields[1:]
    field_names = [f[0] for f in field_defs]

    points: list[dict[str, Any]] = []

    for shape_record in reader.iterShapeRecords():
        shape = shape_record.shape
        record = shape_record.record

        attrs = {field_names[i]: to_json_value(record[i]) for i in range(len(field_names))}

        for xy in shape.points:
            point = {
                "x": float(xy[0]),
                "y": float(xy[1]),
                "attrs": attrs,
            }
            points.append(point)

    xs = [p["x"] for p in points]
    ys = [p["y"] for p in points]

    return {
        "point_count": len(points),
        "fields": field_names,
        "bbox": {
            "min_x": min(xs) if xs else None,
            "max_x": max(xs) if xs else None,
            "min_y": min(ys) if ys else None,
            "max_y": max(ys) if ys else None,
        },
        "points": points,
    }


def load_overview_rows(overview_path: Path) -> tuple[list[dict[str, Any]], str]:
    wb = openpyxl.load_workbook(overview_path, data_only=True, read_only=True)
    sheet_name = "Data" if "Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if not header:
        return [], sheet_name

    headers = [str(h).strip() if h is not None else "" for h in header]
    normalized_headers = [normalize_header(h) for h in headers]

    parsed: list[dict[str, Any]] = []
    for row in rows:
        if row is None:
            continue

        record: dict[str, Any] = {}
        empty = True
        for idx, cell in enumerate(row):
            if idx >= len(headers):
                break
            key = headers[idx]
            norm_key = normalized_headers[idx]
            value = to_json_value(cell)
            record[key] = value
            record[norm_key] = value
            if value not in (None, ""):
                empty = False

        if not empty:
            parsed.append(record)

    return parsed, sheet_name


def downsample_2d(arr: np.ndarray, target_size: int) -> list[list[float | None]]:
    if arr.ndim != 2:
        return []

    rows, cols = arr.shape
    row_idx = np.linspace(0, rows - 1, min(target_size, rows)).astype(int)
    col_idx = np.linspace(0, cols - 1, min(target_size, cols)).astype(int)
    sample = arr[np.ix_(row_idx, col_idx)]

    out: list[list[float | None]] = []
    for r in sample:
        row_out: list[float | None] = []
        for v in r:
            fv = float(v)
            row_out.append(fv if math.isfinite(fv) else None)
        out.append(row_out)
    return out


def safe_minmaxmean(arr: np.ndarray) -> dict[str, float | None]:
    finite = arr[np.isfinite(arr)]
    if finite.size == 0:
        return {"min": None, "max": None, "mean": None}
    return {
        "min": float(np.min(finite)),
        "max": float(np.max(finite)),
        "mean": float(np.mean(finite)),
    }


def load_mat_summary(path: Path, velocity_key: str, mask_key: str, sample_grid_size: int) -> dict[str, Any]:
    mat = sio.loadmat(path, squeeze_me=True, struct_as_record=False)

    result: dict[str, Any] = {
        "source_file": path.name,
        "available_keys": sorted([k for k in mat.keys() if not k.startswith("__")]),
    }

    vel = np.array(mat.get(velocity_key)) if velocity_key in mat else None
    mask = np.array(mat.get(mask_key)) if mask_key in mat else None
    xinterp = np.array(mat.get("xinterp")) if "xinterp" in mat else None
    zinterp = np.array(mat.get("zinterp")) if "zinterp" in mat else None

    if vel is not None and vel.ndim == 2:
        result["velocity"] = {
            "rows": int(vel.shape[0]),
            "cols": int(vel.shape[1]),
            "stats": safe_minmaxmean(vel),
            "sample": downsample_2d(vel, sample_grid_size),
        }
    else:
        result["velocity"] = None

    if mask is not None and mask.ndim == 2:
        mask_float = mask.astype(float)
        result["mask"] = {
            "rows": int(mask.shape[0]),
            "cols": int(mask.shape[1]),
            "water_fraction": float(np.mean(mask_float > 0.5)),
            "sample": downsample_2d(mask_float, sample_grid_size),
        }
    else:
        result["mask"] = None

    if xinterp is not None:
        xinterp = np.array(xinterp).astype(float).flatten()
        result["xinterp"] = [float(v) if math.isfinite(v) else None for v in xinterp]
    else:
        result["xinterp"] = None

    if zinterp is not None:
        zinterp = np.array(zinterp).astype(float).flatten()
        result["zinterp"] = [float(v) if math.isfinite(v) else None for v in zinterp]
        finite = zinterp[np.isfinite(zinterp)]
        result["z_stats"] = {
            "min": float(np.min(finite)) if finite.size else None,
            "max": float(np.max(finite)) if finite.size else None,
            "mean": float(np.mean(finite)) if finite.size else None,
        }
    else:
        result["zinterp"] = None
        result["z_stats"] = None

    return result


def build_cross_sections(
    overview_rows: list[dict[str, Any]],
    mat_summaries: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    by_name = {}
    for row in overview_rows:
        key = row.get("processed_mat_file_name")
        if key:
            by_name[str(key)] = row

    sections: list[dict[str, Any]] = []

    for i, (mat_name, summary) in enumerate(sorted(mat_summaries.items()), start=1):
        row = by_name.get(mat_name, {})

        sx = row.get("cross_section_start_x__utm")
        sy = row.get("cross_section_start_y__utm")
        ex = row.get("cross_section_end_x__utm")
        ey = row.get("cross_section_end_y__utm")

        has_line = all(v is not None for v in [sx, sy, ex, ey])

        center_x = None
        center_y = None
        if has_line:
            center_x = float(sx + ex) / 2.0
            center_y = float(sy + ey) / 2.0

        sections.append(
            {
                "id": i,
                "mat_file": mat_name,
                "transect": row.get("transect"),
                "description": row.get("description"),
                "date": row.get("date"),
                "time_local": row.get("time__local"),
                "start_lat": row.get("start_latitude"),
                "start_lon": row.get("start_longitude"),
                "end_lat": row.get("end_latitude"),
                "end_lon": row.get("end_longitude"),
                "Q_m3s": row.get("q__m3_s"),
                "B_m": row.get("b__m"),
                "T_m": row.get("t__m"),
                "U_ms": row.get("u__m_s"),
                "line": {
                    "start_x": sx,
                    "start_y": sy,
                    "end_x": ex,
                    "end_y": ey,
                    "has_geometry": has_line,
                },
                "center": {
                    "x": center_x,
                    "y": center_y,
                },
                "mat_summary": summary,
            }
        )

    return sections


def main() -> None:
    args = parse_args()

    shp, dbf, shx, overview, mat_files, source, extracted_dir = resolve_inputs(args)

    print(f"Loading shapefile: {shp.name}")
    river_banks = load_shapefile_points(shp, dbf, shx)
    print(f"  points: {river_banks['point_count']}")

    print(f"Loading overview: {overview.name}")
    overview_rows, sheet_name = load_overview_rows(overview)
    print(f"  rows: {len(overview_rows)} (sheet: {sheet_name})")

    sonar_zip_paths, sonar_csv_paths = resolve_sonar_inputs(args, extracted_dir, source)
    elevation_tif_paths = resolve_elevation_inputs(args, extracted_dir, source)
    sonar_bottom = None
    if sonar_zip_paths or sonar_csv_paths:
        print("Loading sonar bottom data")
        if sonar_zip_paths:
            print(f"  sonar zips: {len(sonar_zip_paths)}")
        if sonar_csv_paths:
            print(f"  sonar csv candidates: {len(sonar_csv_paths)}")
        sonar_bottom = load_sonar_bottom_data(
            overview_rows=overview_rows,
            sonar_zip_paths=sonar_zip_paths,
            sonar_csv_paths=sonar_csv_paths,
            max_points=args.sonar_max_points,
            min_depth_m=args.sonar_min_depth_m,
        )
        if sonar_bottom:
            print(
                "  sonar points: "
                f"{sonar_bottom['point_count_sampled']}/{sonar_bottom['point_count_input']} "
                "(sampled/input)"
            )
        else:
            print("  sonar points: none found in provided csv files")

    elevation_raster = None
    if elevation_tif_paths:
        print("Loading elevation raster data")
        print(f"  elevation TIFF candidates: {len(elevation_tif_paths)}")
        elevation_raster = load_elevation_raster_data(
            elevation_tif_paths=elevation_tif_paths,
            max_grid=args.elevation_max_grid,
            river_bbox=river_banks.get("bbox"),
        )
        if elevation_raster:
            sample_meta = elevation_raster.get("sample", {})
            print(
                "  elevation sample grid: "
                f"{sample_meta.get('rows', 0)}x{sample_meta.get('cols', 0)} "
                f"({sample_meta.get('valid_count', 0)} valid cells)"
            )
        else:
            print("  elevation raster: no usable cells found")

    print(f"Loading MAT summaries ({len(mat_files)} files)")
    mat_summaries: dict[str, dict[str, Any]] = {}
    for idx, mat_path in enumerate(mat_files, start=1):
        if idx % 25 == 0 or idx == 1 or idx == len(mat_files):
            print(f"  {idx}/{len(mat_files)}: {mat_path.name}")
        try:
            mat_summaries[mat_path.name] = load_mat_summary(
                mat_path,
                velocity_key=args.velocity_key,
                mask_key=args.mask_key,
                sample_grid_size=args.sample_grid_size,
            )
        except Exception as exc:
            mat_summaries[mat_path.name] = {
                "source_file": mat_path.name,
                "error": str(exc),
            }

    cross_sections = build_cross_sections(overview_rows, mat_summaries)

    package = {
        "river_id": args.river_id,
        "generated_at_utc": dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat(),
        "source": source,
        "overview": {
            "path": str(overview),
            "sheet": sheet_name,
            "row_count": len(overview_rows),
        },
        "elevation_raster": elevation_raster,
        "sonar_bottom": sonar_bottom,
        "river_banks": river_banks,
        "cross_sections": cross_sections,
    }

    out = (
        Path(args.out).expanduser().resolve()
        if args.out
        else (Path("public") / "river-packages" / f"{args.river_id}.json").resolve()
    )
    out.parent.mkdir(parents=True, exist_ok=True)

    with out.open("w", encoding="utf-8") as f:
        json.dump(package, f, ensure_ascii=True, separators=(",", ":"))

    placed = sum(1 for cs in cross_sections if cs["line"]["has_geometry"])
    print(f"Wrote package: {out}")
    print(f"Cross-sections: {len(cross_sections)} total, {placed} with geometry")


if __name__ == "__main__":
    main()
